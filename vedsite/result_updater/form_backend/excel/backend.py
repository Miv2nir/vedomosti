from ..__protocol import FormBackendProtocol
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


def CreateTable(names, path, data={}):
    # C6EFCE - stepik
    # FFEB9C - yandex

    # print(len(names))
    # print(names)

    col = 2
    r = 3
    thins = Side(border_style="thin", color="000000")

    wb = load_workbook(str(path) + "/table.xlsx")
    ws = wb.active
    ws.title = "Контесты"
    # print(wb.sheetnames)
    # ws.column_dimensions['A'].width = 36
    ws["A3"] = "Student"
    ws["A3"].border = Border(top=thins, bottom=thins, right=thins, left=thins)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, len(names) + 1):
        ws.cell(column=1, row=row + 3, value=names[row - 1])
        ws.cell(column=1, row=row + 3, value=names[row - 1]).border = Border(top=thins, bottom=thins, right=thins,
                                                                             left=thins)

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(data["full_name"]) + 2)
    checking_system = ws['B1']
    checking_system.value = data["checking_system_name"]
    checking_system.alignment = Alignment(horizontal="center", vertical="center")
    checking_system.fill = PatternFill("solid", fgColor=data["color"])
    checking_system.font = Font(bold=True, name='Sans', size=24)
    checking_system.border = Border(top=thins, bottom=thins, right=thins, left=thins)

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(data["full_name"]) + 2)
    contest_name = ws['B2']
    contest_name.value = data["contest_title"]
    contest_name.alignment = Alignment(horizontal="center", vertical="center")
    contest_name.font = Font(bold=False, name='Sans', size=22)
    contest_name.border = Border(top=thins, bottom=thins, right=thins, left=thins)
    # ws.freeze_panes = "B1"
    steps = []
    for step in data["full_name"]:
        steps.append(*step.keys())

    # steps = [int(x) for x in steps]
    # print(len(steps))

    for step in range(2, len(data["full_name"]) + 2):
        ws.cell(column=step, row=3, value=steps[step - 2])
        ws.cell(column=step, row=3, value=steps[step - 2]).border = Border(top=thins, bottom=thins, right=thins,
                                                                           left=thins)
    ws.cell(column=len(data["full_name"]) + 2, row=3).value = "Score"
    ws.cell(column=len(data["full_name"]) + 2, row=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(column=len(data["full_name"]) + 2, row=3).border = Border(top=thins, bottom=thins, right=thins, left=thins)

    for i in range(len(names)):
        for j in range(len(data["full_name"])):
            ws.cell(column=j + 2, row=i + 4, value=0)
            ws.cell(column=j + 2, row=i + 4).font = Font(bold=False, name='Sans', size=20)
            ws.cell(column=j + 2, row=i + 4).border = Border(top=thins, bottom=thins, right=thins,
                                                             left=thins)

    # for name in data["full_name"]:
    # print(len(name))
    # print(name)
    # for stud in name.values():
    # print(stud)
    # for full_name in stud:
    # r += 1
    # print(len(stud))
    # ws.cell(column = col, row = r, value = 0)
    # print(full_name, r, cnt)
    # print(col)
    # col += 1
    # r = 3
    # print('здесь')

    cnt = 0
    for name in data["full_name"]:
        for stud in name.values():
            if len(stud) == 0:
                continue
            for n in range(len(names)):
                # print(n, len(names))
                r += 1
                if names[n] == stud[cnt]:
                    ws.cell(column=col, row=r, value=1)
                    ws.cell(column=col, row=r).font = Font(bold=False, name='Sans', size=20)
                    ws.cell(column=col, row=r).border = Border(top=thins, bottom=thins,
                                                               right=thins, left=thins)

                    cnt += 1
            col += 1
            r = 3
            cnt = 0

    for rows in range(len(names)):
        score = 0
        for columns in range(len(data["full_name"])):
            score += ws.cell(row=rows + 4, column=columns + 2).value
        # print(row)
        # print(score)
        ws.cell(row=rows + 4, column=len(data["full_name"]) + 2, value=score)
        ws.cell(row=rows + 4, column=len(data["full_name"]) + 2).font = Font(bold=False, name='Sans', size=20)
        ws.cell(row=rows + 4, column=len(data["full_name"]) + 2).border = Border(top=thins, bottom=thins, right=thins,
                                                                                 left=thins)

    wb.save(str(path) + '/table.xlsx')


class ExcelBackend(FormBackendProtocol):
    def __init__(self):
        super().__init__(None)

    def open(self, path):
        self._filename = path
        self._form = openpyxl.load_workbook(filename=path)
        self._const_form = openpyxl.load_workbook(filename=path,
                                                  data_only=True
                                                  )
        self._sheet = None
        self._val_sheet = None

    def save(self):
        from datetime import datetime
        date = datetime.date(datetime.now())
        filename = str(self._filename)  # "/".join(self._filename.split(".")[:-1]) + "_" + str(date) + ".xlsx"
        self._const_form.close()
        self._form.save(filename)

    def create(self, path):
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()
        del wb
        self.open(path)

    def select_sheet(self, sheet_name):
        self._sheet = self._form[sheet_name]
        self._val_sheet = self._const_form[sheet_name]

    def get_cell_value(self, cell_addr):
        return self._val_sheet[cell_addr].value

    def set_cell_value(self, cell_addr, value):
        self._sheet[cell_addr] = value
        self._val_sheet[cell_addr] = value

    def set_cell_color(self, cell_addr, color):
        c = openpyxl.styles.colors.Color(rgb=color)
        self._sheet[cell_addr].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                                        start_color=c)

    def insert_row(self, ind):
        self._sheet.insert_row(ind)
        self._val_sheet.insert_row(ind)
