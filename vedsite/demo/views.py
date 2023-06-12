from django.shortcuts import render
from django.template.defaulttags import register
from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django import forms
from django.contrib.auth import authenticate
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.models import User
import django_cryptography.fields as crypt
import requests
import json
import pathlib
import pandas as pd
import os
import re
import shutil
import openpyxl
import xlsx2html
from uuid import uuid4
from bs4 import BeautifulSoup
from json import JSONDecodeError
# for logins

from django.contrib.auth import authenticate, login, logout

from .models import *
from .forms import *
import result_updater.checking_system.ya_contest as ya_contest
from .utils import *

import result_updater.checking_system.ya_contest.authorization as yauth
import result_updater.checking_system.ya_contest.fetching as yfetch

from result_updater.form_backend.excel.backend import CreateTable
from result_updater.checking_system.stepik.authorization import InvalidCredentialsException


def index(request):
    latest_question_list = Question.objects.order_by('-pub_date')[:5]
    output = ', '.join([q.question_text for q in latest_question_list])
    return HttpResponse(output)
# Create your views here.


def register_user(request):  # going to roughly repeat the login_user view as to the process is largely similar
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user_login = form.cleaned_data['login']
            try:  # check for existing users
                user = User.objects.get(username=user_login)
                return render(request, 'demo/register.html', {'form': form, 'user_found': True})
            except:
                pass  # no existing user found, proceed
            user_password = form.cleaned_data['password']
            if user_password != form.cleaned_data['password_verify']:  # passwords did not match
                return render(request, 'demo/register.html', {'form': form, 'password_mismatch': True})
            user_email = form.cleaned_data['email']

            if (not user_login) or (not user_email) or (not user_password):  # in the case of whether some credentials have been skipped
                return render(request, 'demo/register.html', {'form': form, 'incomplete_form': True})

            user = User.objects.create_user(user_login, user_email, user_password)
            user.save()  # created the user
            if user is not None:  # login the newly created user
                print('New User Logging in '+user_login)  # for the tests
                login(request, user)
            return HttpResponseRedirect('/')
    else:  # prompt the form
        form = RegisterForm()
        if request.user.is_authenticated:  # if logged in, redirect to the main page
            return HttpResponseRedirect('/')
    return render(request, 'demo/register.html', {'form': form})


def login_user(request):
    if request.method == 'POST':  # look for the imports
        form = AuthForm(request.POST)
        if form.is_valid():
            user_login = form.cleaned_data['login']
            user_password = form.cleaned_data['password']
            user = authenticate(username=user_login, password=user_password)

            if request.user.is_authenticated:  # if logged in, redirect to the main page
                return HttpResponseRedirect('/')

            if user is not None:  # accept the form and login the user
                print('Logging in '+user_login)  # for the tests
                login(request, user)
                return HttpResponseRedirect('/')
            else:  # failed credentials check
                form = AuthForm()
                return render(request, 'demo/login.html', {'form': form, 'failed_login': True})
    else:  # prompt the form
        form = AuthForm()
    return render(request, 'demo/login.html', {'form': form})


def logout_user(request):
    if request.user.is_authenticated:
        print('Logging out '+request.user.username)
        logout(request)
    return HttpResponseRedirect('/login/')


def main_page(request):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    form_logout = LogOutForm()
    return render(request, 'demo/main.html', {'username': request.user, 'form_logout': form_logout})


def account(request):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if request.method == 'POST':  # look for the imports
        form = AuthForm(request.POST)
        if form.is_valid():
            user_login = form.cleaned_data['login']
            user_password = form.cleaned_data['password']
            user = authenticate(username=user_login, password=user_password)


def work(request):

    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    # look up named entries for a logged in user
    lookup = Discipline.objects.filter(d_owner=request.user)

    # for item in lookup:
    # print(item.d_name)

    return render(request, 'demo/work.html', {'username': request.user, 'lookup': lookup})


def work_new(request):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if request.method == 'POST':
        form_new = DisciplineForm(request.POST)
        if form_new.is_valid():
            d = Discipline(d_id=uuid4(), d_name=form_new.cleaned_data['d_name'], d_owner=request.user)
            d.save()
            lookup = Discipline.objects.filter(d_owner=request.user)
            # print(form_new.cleaned_data['d_name'])
            # return render(request, 'demo/work.html', {'username': request.user, 'lookup': lookup})
            return HttpResponseRedirect('/work/')
    # look up named entries for a logged in user
    lookup = Discipline.objects.filter(d_owner=request.user)
    form_new = DisciplineForm()
    return render(request, 'demo/work_new.html', {'username': request.user, 'lookup': lookup, 'form': form_new})


@register.filter  # for being able to send dicts over to templates basically
def get_value(dictionary, key):
    return dictionary.get(key)


def work_manage(request):

    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    # look up named entries for a logged in user
    # lookup = Discipline.objects.filter(d_owner=request.user)
    lookup = Discipline.objects.filter(d_owner=request.user)
    l = []
    name_uuid = {}
    for i in lookup:
        l.append(i.d_name)
        name_uuid[i.d_name] = i.d_id

    if request.method == 'POST':
        forml = DisciplineListForm(request.POST, d_names=l)
        if forml.is_valid():
            for name in l:
                d = Discipline.objects.filter(d_owner=request.user, d_name=name)[0]
                print(d.d_name)
                if forml.cleaned_data['d_og_name_'+name] != d.d_name:
                    d.d_name = forml.cleaned_data['d_og_name_'+name]
                    print(forml.cleaned_data['d_og_name_'+name])
                    d.save()
                # print(forml.cleaned_data['d_og_name_'+i])
            return HttpResponseRedirect('/work/')

    # list of forms for each entry in the database

    forml = DisciplineListForm(d_names=l)
    '''
    for item in lookup:
        prefill = {'d_name': item.d_name}
        form_new = DisciplineListForm(prefill)
        print(prefill)
        formlist.append(form_new)
    '''

    return render(request, 'demo/work_manage.html', {'forml': forml, 'name_uuid': name_uuid})


def work_delete(request, d_id):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if request.method == 'POST':
        del_discipline(d_id, request.user)
        return HttpResponseRedirect('/work/manage/')
    d = Discipline.objects.filter(d_owner=request.user, d_id=d_id)[0]
    lookup = DisciplineGroup.objects.filter(d_id=d_id)
    l = []
    s = []
    for i in lookup:
        l.append(i.g_number)
    slookup = Student.objects.filter(d_id=d_id)
    for i in slookup:
        if i.s_ya_name and i.s_stepik_name:
            query = i.s_ya_name+' - '+i.s_stepik_name
        elif i.s_ya_name:
            query = i.s_ya_name
        else:
            query = i.s_stepik_name
        s.append(query)
    return render(request, 'demo/work_delete.html', {'d_id': d_id, 'd_name': d.d_name, 'g_list': l, 's_list': s})


def discipline(request, d_id):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    # get discipline
    d = Discipline.objects.filter(d_id=d_id)[0]
    lookup = DisciplineGroup.objects.filter(d_id=d_id)
    return render(request, 'demo/group.html', {'username': request.user, 'lookup': lookup, 'd_id': d_id, 'd_name': d.d_name})


def discipline_new(request, d_id):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    d = Discipline.objects.filter(d_id=d_id)[0]
    if request.method == 'POST':
        form_new = GroupForm(request.POST)
        if form_new.is_valid():
            g = DisciplineGroup(g_id=uuid4(), g_number=form_new.cleaned_data['g_number'], d_id=d_id, g_owner=request.user)
            g.save()
            lookup = Discipline.objects.filter(d_owner=request.user)

            # return render(request, 'demo/group.html', {'username': request.user, 'lookup': lookup, 'd_name': d.d_name})
            return HttpResponseRedirect('/work/'+str(d_id)+'/')
    lookup = DisciplineGroup.objects.filter(d_id=d_id)
    form_new = GroupForm()
    return render(request, 'demo/group_new.html', {'username': request.user, 'lookup': lookup, 'd_name': d.d_name, 'd_id': d_id, 'form': form_new})


def discipline_manage(request, d_id):

    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    # look up named entries for a logged in user
    # lookup = Discipline.objects.filter(d_owner=request.user)
    lookup = DisciplineGroup.objects.filter(d_id=d_id)
    l = []
    for i in lookup:
        l.append(i.g_number)

    if request.method == 'POST':
        forml = GroupListForm(request.POST, g_numbers=l)
        if forml.is_valid():
            for num in l:
                g = DisciplineGroup.objects.filter(g_owner=request.user, d_id=d_id, g_number=num)[0]
                print(g.g_number)
                if forml.cleaned_data['g_og_number_'+str(num)] != g.g_number:
                    g.g_number = forml.cleaned_data['g_og_number_'+str(num)]
                    print(forml.cleaned_data['g_og_number_'+str(num)])
                    g.save()
                # print(forml.cleaned_data['d_og_name_'+i])
            return HttpResponseRedirect('/work/'+str(d_id)+'/')
    forml = GroupListForm(g_numbers=l)
    return render(request, 'demo/group_manage.html', {'forml': forml, 'd_id': d_id})


def discipline_delete(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    if request.method == 'POST':
        del_group(g_number, d_id, request.user)
        return HttpResponseRedirect('/work/'+str(d_id)+'/manage/')

    s = []
    slookup = Student.objects.filter(d_id=d_id, g_number=g_number)
    for i in slookup:
        if i.s_ya_name and i.s_stepik_name:
            query = i.s_ya_name+' - '+i.s_stepik_name
        elif i.s_ya_name:
            query = i.s_ya_name
        else:
            query = i.s_stepik_name
        s.append(query)
    return render(request, 'demo/group_delete.html', {'g_number': g_number, 'd_id': d_id, 's_list': s})


def table(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')

    print(pathlib.Path().resolve())
    # find a table dir, make one if absent
    p = pathlib.Path('./generated/'+str(d_id)+'/'+str(g_number))  # dir with tables
    if not p.is_dir():
        p.mkdir(parents=True, exist_ok=True)
    # convert xlsx to html, must be named as table.xlsx
    # print(str(p)+'/table.xlsx')

    # get user groups
    l = student_name_interface(d_id, g_number, mode='all')
    # print(l)

    # gen instance of a table with only the first sheet present
    try:
        table_set_first_sheet(str(p))
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(str(p)+'/table.xlsx')
        table_set_first_sheet(str(p))

    # see if the table is empty
    wb = openpyxl.load_workbook(str(p)+'/table.xlsx')
    sheet = wb.active
    # for cells in sheet.rows:
    #    for cell in cells:
    #        #print(cell.value)

    if [cell.value for cells in sheet.rows for cell in cells] == []:
        return render(request, 'demo/table.html', {'username': request.user, 'd_id': d_id, 'g_number': g_number, 'emptyTable': True, 'l': l})

    # clear empty rows
    table_clear_empty_rows(str(p), table_name='tablecache.xlsx')
    t = xlsx2html.xlsx2html(str(p)+'/tablecache.xlsx')
    t.seek(0)

    # get the body out of the html element
    soup = BeautifulSoup(t.read(), "lxml")
    # return HttpResponse(soup.body.table)
    # print(len(re.findall("\"*!.*3\"", str(soup.body.table))))
    # fix column formatting by adding more spacing arguments
    soup.body.table.colgroup
    coln = len(re.findall("\"*!.*3\"", str(soup.body.table)))
    # replace style="width with style="min-width as it causes rendering issues
    for i in soup.body.table.colgroup.select('col'):
        # soup.body.table.colgroup.insert(3, soup.new_tag(str(i).replace("width", "min-width")[1:-2]))
        # print(str(i).replace("width", "min-width")[1:-2])
        i.extract()
    # regen col styles
    soup.body.table.attrs['id'] = 'xlsxtable'
    soup.body.table.attrs['class'] = 'xlsx'
    # soup.body.table.colgroup.insert(0, soup.new_tag('col style="min-width: 172.8px"'))
    # soup.body.table.colgroup.insert(1, soup.new_tag('col style="min-width: 172.8px"'))
    # soup.body.table.colgroup.insert(1, soup.new_tag('col style="min-width: 120px"'))
    for i in range(coln):
        soup.body.table.colgroup.insert(3, soup.new_tag('col style="min-width: 87.36px"'))

    fileform = UploadFileForm()

    return render(request, 'demo/table.html', {'username': request.user, 'd_id': d_id, 'g_number': g_number, 'emptyTable': False, 'fileform': fileform, 'xlsx': str(soup.body.table), 'l': l})


def table_update(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    table_path = pathlib.Path('./generated/'+str(d_id)+'/'+str(g_number))
    if request.method == 'POST':
        lookup = ImportLog.objects.filter(d_id=d_id, g_number=g_number)
        for i in lookup:
            fill_table(request.user, i.i_type, i.i_contest, d_id=d_id, g_number=g_number, table_path=table_path, col=i.t_col)
    return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/')


def table_download(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    table_path = pathlib.Path('./generated/'+str(d_id)+'/'+str(g_number)+'/table.xlsx')
    # with open(table_path, 'rb') as fh:
    # response = HttpResponse(content_type='application/force-download')
    # response['Content-Disposition'] = 'attachment; filename=%s' % 'table.xlsx'
    # response['Content-Disposition'] = "attachment; filename=%s" % fh
    # response['X-Sendfile'] = smart_str(table_path)
    response = FileResponse(open(table_path, 'rb'))
    return response


def table_upload(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    table_path = pathlib.Path('./generated/'+str(d_id)+'/'+str(g_number)+'/table.xlsx')
    if request.method == 'POST' and request.FILES['table']:
        t = request.FILES['table']
        fs = FileSystemStorage()
        if os.path.isfile(table_path):
            os.remove(table_path)
        fs.save(table_path, t)

        # handle t_col
        wb = openpyxl.load_workbook(table_path)
        ws = wb.active
        lookup = DisciplineGroup.objects.filter(d_id=d_id, g_number=g_number)[0]
        lookup.t_col = ws.max_column+2
        lookup.save()

        # form = UploadFileForm(request.POST, request.FILES)
        # print(form.errors)
        # if form.is_valid():
        #    print("HIIIIIIIIIII")
#
        #    with open(table_path, 'wb+') as t:
        #        for chunk in request.FILES["file"].chunks():
        #            t.write(chunk)
        return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/')
    fileform = UploadFileForm()
    return render(request, 'demo/table_upload.html', {'d_id': d_id, 'g_number': g_number, 'table_path': table_path, 'fileform': fileform})


def table_delete(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
        # verify if user is owner of the object being looked up
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    if request.method == 'POST':
        l = ImportLog.objects.filter(d_id=d_id, g_number=g_number)
        for n in l:
            n.delete()
        # wipe the table
        shutil.rmtree('./generated/'+str(d_id)+'/'+str(g_number)+'')
        group = DisciplineGroup.objects.filter(g_owner=request.user, d_id=d_id, g_number=g_number)[0]
        group.t_col = 2
        group.save()
        return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/')
    return render(request, 'demo/table_delete.html', {'g_number': g_number, 'd_id': d_id, 'g_number': g_number})


def imports(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    if request.method == 'POST':
        print('post')
        form = PlatformSelectForm(request.POST)
        if form.is_valid():
            # print(form.cleaned_data['p_choice'])
            # print(form.cleaned_data['p_id'])
            user = request.user
            source = form.cleaned_data['p_choice']
            task_id = form.cleaned_data['p_id']
            table_path = pathlib.Path('./generated/'+str(d_id)+'/'+str(g_number))
            group = DisciplineGroup.objects.filter(d_id=d_id, g_number=g_number, g_owner=request.user)[0]

            col = group.t_col
            ilog = ImportLog(i_id=uuid4(), d_id=d_id, g_number=g_number, i_type=source, i_contest=task_id, t_col=col)
            print(col)
            try:
                col = fill_table(user, source, task_id, d_id=d_id, g_number=g_number, table_path=table_path, col=col)+2
            except KeyError:
                form = PlatformSelectForm()
                return render(request, 'demo/imports.html', {'username': request.user, 'd_id': d_id, 'g_number': g_number, 'form': form, 'keyError': True})
            except (JSONDecodeError, InvalidCredentialsException, IndexError):
                return render(request, 'demo/imports.html', {'username': request.user, 'd_id': d_id, 'g_number': g_number, 'form': form, 'jsonError': True})
            print(col)
            group.t_col = col
            ilog.save()
            group.save()
            return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/')
    form = PlatformSelectForm()
    return render(request, 'demo/imports.html', {'username': request.user, 'd_id': d_id, 'g_number': g_number, 'form': form})


def student(request, d_id, g_number):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    l = student_name_interface(d_id, g_number, mode='all')
    print(l)
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            if form.cleaned_data['s_ya_name']:
                lookup = Student.objects.filter(d_id=d_id, g_number=g_number, s_ya_name=form.cleaned_data['s_ya_name'])
            elif form.cleaned_data['s_stepik_name']:
                lookup = Student.objects.filter(d_id=d_id, g_number=g_number, s_stepik_name=form.cleaned_data['s_stepik_name'])

            if lookup:
                st = lookup[0]
            else:
                st = Student(s_id=uuid4(), s_owner=request.user, d_id=d_id, g_number=g_number)
                # st.s_id = lookup[0].s_id
            # st.s_display_name = form.cleaned_data['s_display_name']
            # st.s_email = form.cleaned_data['s_email']
            st.s_ya_name = form.cleaned_data['s_ya_name']
            st.s_stepik_name = form.cleaned_data['s_stepik_name']
            st.save()

        return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/students/')
    form = StudentForm()

    return render(request, 'demo/students.html', {'username': request.user, 'form': form, 'd_id': d_id, 'g_number': g_number, 'l': l})


def student_delete(request, d_id, g_number, s_id):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    if not Discipline.objects.filter(d_owner=request.user, d_id=d_id):
        return HttpResponseRedirect('/work/')
    if request.method == 'POST':
        st = Student.objects.filter(d_id=d_id, g_number=g_number, s_id=s_id)[0]
        st.delete()
        # print('yeet', st)
    return HttpResponseRedirect('/work/'+str(d_id)+'/'+str(g_number)+'/students/')


def account(request):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    user = User.objects.get(username=request.user)
    if request.method == 'POST':
        # validate retrieved password
        form = NewPasswordForm(request.POST)
        if form.is_valid():
            if user.check_password(form.cleaned_data['password_old']) and form.cleaned_data['password'] == form.cleaned_data['password']:
                user.set_password(form.cleaned_data['password'])
                user.save()
                return render(request, 'demo/account.html', {'username': request.user, 'form': form, 'change_success': True})
            else:
                return render(request, 'demo/account.html', {'username': request.user, 'form': form, 'change_fail': True})

    form = NewPasswordForm()
    return render(request, 'demo/account.html', {'username': request.user, 'form': form})


def credentials(request):
    if not request.user.is_authenticated:  # user not yet logged in
        return HttpResponseRedirect('/login/')
    # creds = Credentials.objects.filter(user=request.user)
    p = pathlib.Path('./generated/cookies')
    lookup = Credentials.objects.filter(user_name=request.user)
    if request.method == 'POST':
        # validate retrieved password
        form = CredentialsForm(request.POST)
        if form.is_valid():
            if not lookup:
                creds = Credentials(user_name=request.user)
            else:
                creds = lookup[0]

            if form.cleaned_data['ya_l'] and form.cleaned_data['ya_p']:
                # print('YANDEX!!!')
                d = {"login": form.cleaned_data['ya_l'], 'password': form.cleaned_data['ya_p']}
                yuser = yauth.YaContestAuthorizer(authorize=False)
                yuser.authorize(d)
                if not os.path.exists('./generated/cookies/'):
                    os.makedirs('./generated/cookies/')
                write_cookies(p, request.user.username, yuser._session)
                creds.ya_login = form.cleaned_data['ya_l']
            if form.cleaned_data['step_api'] and form.cleaned_data['step_id']:
                creds.stepik_id = form.cleaned_data['step_id']
                creds.stepik_key = form.cleaned_data['step_api']
            creds.save()

            return HttpResponseRedirect('/credentials/')
            # creds=Credentials

    form = CredentialsForm()
    isLookup = False
    ya_l = step_api = step_id = ''
    if lookup:
        isLookup = True
        ya_l = lookup[0].ya_login
        step_id = lookup[0].stepik_id
        # step_api = lookup[0].stepik_id[:5]+'...'
    return render(request, 'demo/credentials.html', {'user': request.user, 'form': form, 'isLookup': isLookup, 'isStepik': step_api != '...', 'ya_l': ya_l, 'step_id': step_id})


def credentials_about(request):
    return render(request, 'demo/credentials_about.html')


'''
def credentials(request):  # /credentials page
    print("Retrieving some stuff")
    if request.method == 'POST':
        form_credentials = CredentialsForm(request.POST)
        if form_credentials.is_valid():
            # look up the corresponding teacher in database

            ya_login = form_credentials.cleaned_data['ya_login']
            ya_password = form_credentials.cleaned_data['ya_password']
            ya_id = form_credentials.cleaned_data['ya_id']
            # print(ya_login)  # haha leaking personal info

            ya_user = ya_contest.authorization.YaContestAuthorizer(authorize=False)  # create an authorizer object
            ya_user.authorize({"login": ya_login, "password": ya_password})  # authorize the user (should refresh the session hopefully)
            if not Teacher.objects.filter(linked_user=request.user.username):  # check if a teacher object exists in a DB already
                ya_object = Teacher(linked_user=request.user.username, yandex_login=ya_login)  # creates a Teacher object in case of absense
            else:
                ya_object = Teacher.objects.filter(linked_user=request.user.username)
            ya_object.ya_session = ya_user.get_session()
            # print(ya_object.ya_session)
            # ya_object.save()
            # apparently am unable to save the session so gonna retrieve everything in one go now
            filename = request.user.username+'_'+str(ya_id)+'.json'
            pth = str(pathlib.Path(__file__).parent)
            ya_get = ya_contest.fetching.YaContestUpdFetcher(ya_user.get_session(), {ya_id: [ya_id]}, "automatic", pth+'/static/generated/'+filename)
            ya_get.fetch_updates()

            with open(pth+'/static/generated/'+filename, 'r', encoding='utf8') as f:
                try:
                    j = _digest(json.load(f), ya_id)  # turns retrieved .json into a desired format
                except:  # incorrect json means server did not tolerate the request
                    return render(request, 'demo/credentials.html', {'form_credentials': form_credentials, 'changes_saved': False, 'relog_needed': False, 'crash': True})
            with open(pth+'/static/generated/processed_'+filename, 'w', encoding='utf8') as f:
                f.write(json.dumps(j, ensure_ascii=False))
                print(pth+'/static/generated/processed_'+filename)
                os.remove(pth+'/static/generated/'+filename)  # no longer needed
            _json_to_csv(pth+'/static/generated/processed_'+filename, filename[:-5]+'.csv', pth)  # make a csv variant of the data
            with open(pth+'/static/generated/'+filename[:-5]+'.csv', 'r', encoding='utf8') as f:  # fixing up the csv lol
                deeta = f.readlines()
                deeta[0] = 'contest_students,contest_ID,results\n'
            with open(pth+'/static/generated/'+filename[:-5]+'.csv', 'w', encoding='utf8') as f:
                f.writelines(deeta)
            # purge old files
            purgetemplate = ['processed_'+filename, filename[:-5]+'.csv']
            print("current filename is ", filename)
            for purgename in os.listdir(pth+'/static/generated/'):
                # (purgename in purgetemplate) != (request.user.username in purgename)
                print(purgetemplate, purgename, ya_login, (request.user.username in purgename) and not (str(ya_id) in purgename))
                if (request.user.username in purgename) and not (str(ya_id) in purgename):
                    os.remove(pth+'/static/generated/'+purgename)

            return render(request, 'demo/credentials.html', {'form_credentials': form_credentials, 'changes_saved': True, 'relog_needed': False, 'pathcsv': ('/static/generated/'+filename[:-5]+'.csv'), 'pathjson': ('/static/generated/processed_'+filename)})
        else:
            form_credentials = CredentialsForm()
            return render(request, 'demo/credentials.html', {'form_credentials': form_credentials, 'changes_saved': False, 'relog_needed': False})
    else:
        form_credentials = CredentialsForm()
        return render(request, 'demo/credentials.html', {'form_credentials': form_credentials, 'changes_saved': False, 'relog_needed': False})
'''
